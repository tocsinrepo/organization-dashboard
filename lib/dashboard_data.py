"""
Read an organization's *real* Income / Expense / Data numbers out of an
uploaded dashboard workbook, so the live preview can show that organization's
actual figures instead of the illustrative placeholder numbers.

This reads the same standard template that excel_writer.py documents. On the
"Dashboard" sheet the charts are, in order:
  [0] LineChart "Income"   - up to 4 series (fiscal years)
  [1] LineChart "Expense"  - up to 4 series (fiscal years)
  [2] BarChart  "Data"     - series 0 = Actual, series 1 = Budgeted

Each chart series' values are stored as a *reference* to a range on another
sheet (e.g. "Summary!$C$5:$C$16"), not as inline numbers. We resolve those
references against the workbook's cached cell values (loaded with
data_only=True, which returns the last value Excel computed and saved) to get
the actual numbers. Files that have been opened and saved in Excel carry those
cached values; a freshly generated file that has never been opened may not, in
which case we fall back gracefully (see extract_dashboard_data's `ok` flag).

Nothing here writes to the workbook. Financial data is read-only and never
modified — consistent with excel_writer.py leaving Summary/Raw untouched.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries

DASHBOARD_SHEET = "Dashboard"


@dataclass
class Series:
    name: str
    values: list[float] = field(default_factory=list)


@dataclass
class DashboardData:
    categories: list = field(default_factory=list)   # shared x labels (e.g. months)
    income: list[Series] = field(default_factory=list)
    expense: list[Series] = field(default_factory=list)
    bar_categories: list = field(default_factory=list)   # e.g. ["Actual", "Budgeted"]
    bar_values: list = field(default_factory=list)       # numbers aligned to bar_categories
    ok: bool = False        # True only if we recovered at least some real numbers
    note: str = ""          # human-readable reason when ok is False


def _split_ref(formula: str):
    """Split a chart reference like "Summary!$C$5:$C$16" or "'My Sheet'!$B$2"
    into (sheet_name, a1_range). Returns (None, None) if it can't be parsed."""
    if not formula:
        return None, None
    f = formula.lstrip("=").strip()
    # Reference may point into another workbook: [1]Sheet!... — not resolvable here.
    if f.startswith("["):
        return None, None
    if "!" not in f:
        return None, None
    sheet_part, range_part = f.rsplit("!", 1)
    sheet = sheet_part.strip()
    if sheet.startswith("'") and sheet.endswith("'"):
        sheet = sheet[1:-1].replace("''", "'")
    return sheet, range_part.replace("$", "")


def _resolve_values(wb, formula) -> list:
    """Return the cached numeric values in the range a chart series points at."""
    sheet, rng = _split_ref(formula)
    if sheet is None or sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    try:
        min_col, min_row, max_col, max_row = range_boundaries(rng)
    except Exception:
        return []
    out = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            out.append(cell.value)
    return out


def _resolve_scalar(wb, formula):
    """Return a single cached value (used for a series' name reference)."""
    vals = _resolve_values(wb, formula)
    for v in vals:
        if v is not None:
            return v
    return None


def _series_name(wb, ser, fallback: str) -> str:
    tx = getattr(ser, "tx", None)
    if tx is not None:
        # Named reference (strRef) or an inline literal (v)
        ref = getattr(tx, "strRef", None) or getattr(tx, "numRef", None)
        if ref is not None and getattr(ref, "f", None):
            val = _resolve_scalar(wb, ref.f)
            if val not in (None, ""):
                return str(val)
        v = getattr(tx, "v", None)
        if v not in (None, ""):
            return str(v)
    return fallback


def _series_values(wb, ser) -> list[float]:
    val = getattr(ser, "val", None)
    ref = getattr(val, "numRef", None) if val is not None else None
    if ref is not None and getattr(ref, "f", None):
        raw = _resolve_values(wb, ref.f)
        return [float(x) if isinstance(x, (int, float)) else None for x in raw]
    return []


def _series_categories(wb, ser) -> list:
    cat = getattr(ser, "cat", None)
    if cat is not None:
        ref = getattr(cat, "strRef", None) or getattr(cat, "numRef", None)
        if ref is not None and getattr(ref, "f", None):
            return [x for x in _resolve_values(wb, ref.f)]
    return []


def _line_series(wb, chart, kind: str) -> list[Series]:
    out = []
    for i, ser in enumerate(chart.series[:4]):
        name = _series_name(wb, ser, f"{kind} {i + 1}")
        values = _series_values(wb, ser)
        out.append(Series(name=name, values=values))
    return out


def extract_dashboard_data(source) -> DashboardData:
    """
    Read real Income/Expense/Data numbers from a dashboard workbook.

    `source` may be a path, bytes, or a file-like object (e.g. a Streamlit
    UploadedFile). Never raises for a bad/foreign file — returns a DashboardData
    with ok=False and a `note` explaining why, so the caller can quietly fall
    back to placeholder preview data.
    """
    try:
        if isinstance(source, (str, Path)):
            wb = openpyxl.load_workbook(source, data_only=True)
        elif isinstance(source, (bytes, bytearray)):
            wb = openpyxl.load_workbook(BytesIO(bytes(source)), data_only=True)
        else:
            data = source.getvalue() if hasattr(source, "getvalue") else source.read()
            wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    except Exception as e:  # not a valid xlsx, etc.
        return DashboardData(ok=False, note=f"Could not open workbook: {e}")

    if DASHBOARD_SHEET not in wb.sheetnames:
        return DashboardData(ok=False, note="No 'Dashboard' sheet found.")
    ws = wb[DASHBOARD_SHEET]

    charts = list(ws._charts)
    line_charts = [c for c in charts if type(c).__name__ == "LineChart"]
    bar_charts = [c for c in charts if type(c).__name__ == "BarChart"]

    data = DashboardData()

    if len(line_charts) >= 1:
        data.income = _line_series(wb, line_charts[0], "Income")
        cats = _series_categories(wb, line_charts[0].series[0]) if line_charts[0].series else []
        data.categories = cats
    if len(line_charts) >= 2:
        data.expense = _line_series(wb, line_charts[1], "Expense")

    if bar_charts:
        bar = bar_charts[0]
        vals, names = [], []
        for i, ser in enumerate(bar.series[:2]):
            names.append(_series_name(wb, ser, ["Actual", "Budgeted"][i] if i < 2 else f"Bar {i}"))
            sv = _series_values(wb, ser)
            # A budget-vs-actual bar is usually one number per series.
            nums = [x for x in sv if isinstance(x, (int, float))]
            vals.append(nums[-1] if nums else None)
        data.bar_categories = names
        data.bar_values = vals

    # We consider the extract successful if we recovered any real numbers at all.
    def _has_numbers(series_list):
        return any(any(v is not None for v in s.values) for s in series_list)

    if _has_numbers(data.income) or _has_numbers(data.expense) or any(
        v is not None for v in data.bar_values
    ):
        data.ok = True
    else:
        data.note = (
            "The workbook matched the template but had no saved chart values to "
            "read (open and save it once in Excel so the numbers are cached)."
        )
    return data
