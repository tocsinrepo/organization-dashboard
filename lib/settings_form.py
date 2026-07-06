"""
Downloadable/uploadable Excel "settings form" -- an alternative way to set
an organization's identity (name, header text), colors, fiscal-year display
order, and axis minimums, for people who'd rather fill in an Excel sheet
than click through the web page's controls. Both paths feed the same
OrgProfile fields and lead to the same result.

Layout (fixed cell addresses, shared between build_settings_form and
parse_settings_form so they can never drift apart):

  Organization info
    Organization name                C6
    Header subtitle                  C7

  Color scheme picker
    Preset dropdown                  C10  (one of the 5 names in SCHEMES,
                                            below -- picking one fills in
                                            every color cell below via an
                                            =INDEX/MATCH formula against the
                                            hidden "Schemes" sheet, and each
                                            cell's fill color updates too via
                                            conditional formatting)

  Colors (hex) -- auto-filled by the scheme above; edit any cell directly
  to override just that one color and break its link to the preset.
    Banner - primary                 C14
    Banner - secondary                C15
    Accent strip                     C16
    Bar chart - Actual                C17
    Bar chart - Budgeted             C18
    Line series 1                    C19
    Line series 2                    C20
    Line series 3                    C21
    Line series 4                    C22

  Fiscal year display order (1-4, must be the 4 numbers 1-4 with no repeats)
    Line series 1 -> position        C26
    Line series 2 -> position        C27
    Line series 3 -> position        C28
    Line series 4 -> position        C29

  Axis minimums ($)
    Data bar chart                    C32
    Income/Expense line charts       C33

  Chart types ("line" or "bar"; blank keeps the default)
    Income chart                     C36
    Expense chart                    C37
    Data chart                       C38
"""

from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from lib.org_profile import (
    OrgProfile,
    DEFAULT_BANNER_PRIMARY, DEFAULT_BANNER_SECONDARY, DEFAULT_ACCENT,
    DEFAULT_LINE_COLORS, DEFAULT_BAR_ACTUAL, DEFAULT_BAR_BUDGET,
    GOLD_TEAL_BANNER_PRIMARY, GOLD_TEAL_BANNER_SECONDARY,
    GOLD_TEAL_ACCENT, GOLD_TEAL_LINE_COLORS,
    GOLD_TEAL_BAR_ACTUAL, GOLD_TEAL_BAR_BUDGET,
)

SHEET_NAME = "Settings"
SCHEMES_SHEET_NAME = "Schemes"

# Five preset color schemes. Field names match OrgProfile attributes so a
# scheme dict can be splatted straight onto a profile. "Purple & Orange
# (Full Send style)" is the default -- both here and as OrgProfile's own
# DEFAULT_* constants (imported above, not re-typed) so a brand-new org's
# web-app preview and this form's dropdown default can never disagree with
# each other again.
#
# RULE for line_colors (confirmed by Jon 2026-07-06): line_colors[3] -- the
# 4th line series, which is the MOST CURRENT fiscal year -- must be the
# scheme's DOMINANT color, so the current year visually stands out. The other
# three (line_colors[0..2], older years) are a light->dark gradient of the
# supporting family. Dominant means the contrasting accent where the scheme
# has one (orange for Purple & Orange, red for Slate & Crimson, gold for Gold
# & Teal) and the darkest primary shade for the monochromatic blues (darkest
# royal blue for Blue & Sky, darkest navy for Navy & Silver). Any new scheme
# added here must follow the same rule.
SCHEMES: dict[str, dict] = {
    "Purple & Orange (Full Send style)": {
        "banner_primary": DEFAULT_BANNER_PRIMARY,
        "banner_secondary": DEFAULT_BANNER_SECONDARY,
        "accent": DEFAULT_ACCENT,
        "bar_actual_color": DEFAULT_BAR_ACTUAL,
        "bar_budget_color": DEFAULT_BAR_BUDGET,
        "line_colors": list(DEFAULT_LINE_COLORS),
    },
    "Gold & Teal": {
        "banner_primary": GOLD_TEAL_BANNER_PRIMARY,
        "banner_secondary": GOLD_TEAL_BANNER_SECONDARY,
        "accent": GOLD_TEAL_ACCENT,
        "bar_actual_color": GOLD_TEAL_BAR_ACTUAL,
        "bar_budget_color": GOLD_TEAL_BAR_BUDGET,
        "line_colors": list(GOLD_TEAL_LINE_COLORS),
    },
    "Navy & Silver": {
        "banner_primary": "#1B3A5C",
        "banner_secondary": "#102A44",
        "accent": "#4A90D9",
        "bar_actual_color": "#4A90D9",
        "bar_budget_color": "#C7CDD1",
        # Series 4 (current year) = darkest navy (dominant); older years lighter.
        "line_colors": ["#AFC6DA", "#6699BF", "#4A90D9", "#1B3A5C"],
    },
    "Blue & Sky": {
        "banner_primary": "#1D4ED8",
        "banner_secondary": "#14268C",
        "accent": "#38BDF8",
        "bar_actual_color": "#38BDF8",
        "bar_budget_color": "#D2D2D7",
        # Series 4 (current year) = darkest royal blue (dominant); older years lighter.
        "line_colors": ["#BFDBFE", "#38BDF8", "#60A5FA", "#1D4ED8"],
    },
    "Slate & Crimson": {
        "banner_primary": "#3A4750",
        "banner_secondary": "#262E34",
        "accent": "#C0392B",
        "bar_actual_color": "#C0392B",
        "bar_budget_color": "#D2D2D7",
        "line_colors": ["#B9C2C7", "#74838C", "#3A4750", "#C0392B"],
    },
}
DEFAULT_SCHEME_NAME = "Purple & Orange (Full Send style)"
SCHEME_ORDER = list(SCHEMES.keys())  # fixed order used for the Schemes reference sheet

ORG_FIELD_CELLS = {
    "org_name": ("B6", "C6", "Organization name"),
    "header_subtitle": ("B7", "C7", "Header subtitle"),
}

SCHEME_PICKER_CELL = "C10"

# field name -> (schemes-sheet column letter, settings-sheet color cell, label)
COLOR_CELLS = {
    "banner_primary": ("B", "C14", "Banner - primary"),
    "banner_secondary": ("C", "C15", "Banner - secondary"),
    "accent": ("D", "C16", "Accent strip"),
    "bar_actual_color": ("E", "C17", "Bar chart - Actual"),
    "bar_budget_color": ("F", "C18", "Bar chart - Budgeted"),
}
LINE_COLOR_CELLS = ["C19", "C20", "C21", "C22"]  # line_colors[0..3]
LINE_COLOR_SCHEME_COLS = ["G", "H", "I", "J"]  # matching columns on the Schemes sheet

DISPLAY_ORDER_CELLS = ["C26", "C27", "C28", "C29"]  # display_order slot for series 1..4
BAR_AXIS_MIN_CELL = "C32"
LINE_AXIS_MIN_CELL = "C33"

# Chart type per chart ("line" or "bar"). Blank on upload = keep the field's
# template default rather than erroring.
CHART_TYPE_CELLS = {
    "income_chart_type": ("C36", "Income chart", "line"),
    "expense_chart_type": ("C37", "Expense chart", "line"),
    "data_chart_type": ("C38", "Data chart", "bar"),
}
VALID_CHART_TYPES = ("line", "bar")


class SettingsFormError(ValueError):
    """Raised when an uploaded settings form has missing or invalid values."""


def _build_schemes_sheet(wb: openpyxl.Workbook) -> None:
    """Hidden reference sheet the Settings tab's color formulas look up into."""
    ws = wb.create_sheet(SCHEMES_SHEET_NAME)
    headers = ["Scheme name", "banner_primary", "banner_secondary", "accent",
               "bar_actual_color", "bar_budget_color",
               "line1", "line2", "line3", "line4"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, name in enumerate(SCHEME_ORDER, start=2):
        scheme = SCHEMES[name]
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=scheme["banner_primary"])
        ws.cell(row=row_idx, column=3, value=scheme["banner_secondary"])
        ws.cell(row=row_idx, column=4, value=scheme["accent"])
        ws.cell(row=row_idx, column=5, value=scheme["bar_actual_color"])
        ws.cell(row=row_idx, column=6, value=scheme["bar_budget_color"])
        for i, line_hex in enumerate(scheme["line_colors"]):
            ws.cell(row=row_idx, column=7 + i, value=line_hex)

    ws.sheet_state = "hidden"


def _scheme_lookup_formula(scheme_col_letter: str) -> str:
    last_row = 1 + len(SCHEME_ORDER)
    picker_col = SCHEME_PICKER_CELL[0]
    picker_row = SCHEME_PICKER_CELL[1:]
    return (
        f"=INDEX({SCHEMES_SHEET_NAME}!${scheme_col_letter}$2:${scheme_col_letter}${last_row}, "
        f"MATCH(${picker_col}${picker_row}, "
        f"{SCHEMES_SHEET_NAME}!$A$2:$A${last_row}, 0))"
    )


def _add_scheme_swatch_formatting(ws, cell: str, hex_values: list[str]) -> None:
    """Make the cell's own fill match whichever scheme value it currently
    equals, so the swatch stays visually correct as the dropdown changes and
    Excel recalculates -- no VBA needed."""
    for hex_value in hex_values:
        fill = PatternFill(start_color="FF" + hex_value.lstrip("#"),
                            end_color="FF" + hex_value.lstrip("#"), fill_type="solid")
        rule = FormulaRule(formula=[f'EXACT({cell},"{hex_value}")'], fill=fill, stopIfTrue=True)
        ws.conditional_formatting.add(cell, rule)


def build_settings_form(profile: OrgProfile) -> bytes:
    """Return .xlsx file bytes: a filled-in-with-current-values settings form."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 28

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)

    ws["B2"] = "Organization Dashboard Styler -- Settings Form"
    ws["B2"].font = title_font
    ws["B3"] = ("Fill in the values below, save this file, then upload it in the app "
                "under \"Upload a filled-in settings form.\"")
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("B3:D3")
    ws.row_dimensions[3].height = 30

    # --- Organization info ---
    ws["B5"] = "ORGANIZATION INFO"
    ws["B5"].font = bold
    for field, (label_cell, value_cell, label) in ORG_FIELD_CELLS.items():
        ws[label_cell] = label
        ws[value_cell] = getattr(profile, field)

    # --- Color scheme picker ---
    ws["B9"] = "COLOR SCHEME"
    ws["B9"].font = bold
    ws["B10"] = "Pick a preset scheme"
    ws[SCHEME_PICKER_CELL] = DEFAULT_SCHEME_NAME
    ws["B11"] = ("Choosing a preset fills in every color below automatically. "
                 "You can still type over any individual color afterward to fine-tune it.")
    ws["B11"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("B11:D11")
    ws.row_dimensions[11].height = 30

    dv_scheme = DataValidation(
        type="list",
        formula1=f"={SCHEMES_SHEET_NAME}!$A$2:$A${1 + len(SCHEME_ORDER)}",
        showErrorMessage=True, errorTitle="Invalid scheme",
        error="Pick one of the 5 listed color schemes.",
    )
    ws.add_data_validation(dv_scheme)
    dv_scheme.add(ws[SCHEME_PICKER_CELL])

    # --- Colors (formulas referencing the hidden Schemes sheet) ---
    ws["B13"] = "COLORS (hex) -- auto-filled by the scheme above; edit directly to override"
    ws["B13"].font = bold
    ws["B14"] = "Banner - primary"
    ws["B15"] = "Banner - secondary"
    ws["B16"] = "Accent strip"
    ws["B17"] = "Bar chart - Actual"
    ws["B18"] = "Bar chart - Budgeted"

    for field, (scheme_col, value_cell, _label) in COLOR_CELLS.items():
        ws[value_cell] = _scheme_lookup_formula(scheme_col)
        hex_values = [SCHEMES[name][field] for name in SCHEME_ORDER]
        _add_scheme_swatch_formatting(ws, value_cell, hex_values)

    for i, cell in enumerate(LINE_COLOR_CELLS):
        ws[f"B{19 + i}"] = f"Line series {i + 1}"
        ws[cell] = _scheme_lookup_formula(LINE_COLOR_SCHEME_COLS[i])
        hex_values = [SCHEMES[name]["line_colors"][i] for name in SCHEME_ORDER]
        _add_scheme_swatch_formatting(ws, cell, hex_values)

    # --- Fiscal year display order ---
    ws["B24"] = "FISCAL YEAR DISPLAY ORDER (1-4, no repeats)"
    ws["B24"].font = bold
    ws["B25"] = "1 = listed first in the legend AND drawn on top of the other lines."
    ws.merge_cells("B25:D25")

    display_order = profile.display_order or [0, 1, 2, 3]
    positions_by_series = [0, 0, 0, 0]
    for slot, series_idx in enumerate(display_order):
        if 0 <= series_idx < 4:
            positions_by_series[series_idx] = slot + 1

    for i, cell in enumerate(DISPLAY_ORDER_CELLS):
        ws[f"B{26 + i}"] = f"Line series {i + 1} -> display position"
        ws[cell] = positions_by_series[i] or (i + 1)

    dv_order = DataValidation(type="whole", operator="between", formula1=1, formula2=4,
                               showErrorMessage=True, errorTitle="Invalid position",
                               error="Enter a whole number from 1 to 4.")
    ws.add_data_validation(dv_order)
    for cell in DISPLAY_ORDER_CELLS:
        dv_order.add(ws[cell])

    # --- Axis minimums ---
    ws["B31"] = "AXIS MINIMUMS ($)"
    ws["B31"].font = bold
    ws["B32"] = "Data bar chart"
    ws[BAR_AXIS_MIN_CELL] = profile.bar_axis_min
    ws["B33"] = "Income/Expense line charts"
    ws[LINE_AXIS_MIN_CELL] = profile.line_axis_min

    # --- Chart types ---
    ws["B35"] = "CHART TYPES (line or bar)"
    ws["B35"].font = bold
    dv_type = DataValidation(
        type="list", formula1='"line,bar"',
        showErrorMessage=True, errorTitle="Invalid chart type",
        error='Enter "line" or "bar".',
    )
    ws.add_data_validation(dv_type)
    for field, (cell, label, _default) in CHART_TYPE_CELLS.items():
        row = cell[1:]
        ws[f"B{row}"] = label
        ws[cell] = getattr(profile, field)
        dv_type.add(ws[cell])

    _build_schemes_sheet(wb)
    wb.active = wb.sheetnames.index(SHEET_NAME)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_settings_form(file_like) -> dict:
    """
    Read a filled-in settings form (path or file-like/bytes buffer) and return
    a dict of OrgProfile field updates: org name/subtitle (if filled in),
    colors, display_order, axis minimums.

    Raises SettingsFormError with a plain-English message if a required value
    is missing/invalid, or the display-order column isn't a clean 1-4
    permutation. Color cells are formulas -- if this file was never opened
    and saved in Excel/Google Sheets since download, they won't have a
    calculated value yet, which is called out explicitly below.
    """
    wb = openpyxl.load_workbook(file_like, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SettingsFormError(
            f"This doesn't look like a settings form -- no '{SHEET_NAME}' tab found."
        )
    ws = wb[SHEET_NAME]

    updates: dict = {}

    # Organization info is optional -- leave the current profile's values
    # alone if the user didn't fill these in.
    for field, (_, value_cell, _label) in ORG_FIELD_CELLS.items():
        value = ws[value_cell].value
        if value is not None and str(value).strip():
            updates[field] = str(value).strip()

    def _read_color(cell: str, label: str) -> str:
        value = ws[cell].value
        if value is None or not str(value).strip():
            raise SettingsFormError(
                f"'{label}' is empty. If you just downloaded this file, open it in Excel "
                "or Google Sheets and save it first so the color-scheme formulas can "
                "calculate, then upload that saved copy."
            )
        return _normalize_hex(str(value), label)

    for field, (_, value_cell, label) in COLOR_CELLS.items():
        updates[field] = _read_color(value_cell, label)

    line_colors = []
    for i, cell in enumerate(LINE_COLOR_CELLS):
        line_colors.append(_read_color(cell, f"Line series {i + 1}"))
    updates["line_colors"] = line_colors

    positions_by_series = []
    for i, cell in enumerate(DISPLAY_ORDER_CELLS):
        value = ws[cell].value
        try:
            pos = int(value)
        except (TypeError, ValueError):
            raise SettingsFormError(
                f"'Line series {i + 1} -> display position' must be a whole number from 1 to 4."
            )
        positions_by_series.append(pos)

    if sorted(positions_by_series) != [1, 2, 3, 4]:
        raise SettingsFormError(
            "The 4 display-position values must be exactly 1, 2, 3, and 4, each used once. "
            f"Found: {positions_by_series}."
        )

    display_order = [0, 0, 0, 0]
    for series_idx, pos in enumerate(positions_by_series):
        display_order[pos - 1] = series_idx
    updates["display_order"] = display_order

    for field, cell in (("bar_axis_min", BAR_AXIS_MIN_CELL), ("line_axis_min", LINE_AXIS_MIN_CELL)):
        value = ws[cell].value
        try:
            updates[field] = float(value)
        except (TypeError, ValueError):
            raise SettingsFormError(f"Axis minimum in cell {cell} must be a number.")

    # Chart types are optional -- a blank cell keeps the field's default.
    for field, (cell, label, default) in CHART_TYPE_CELLS.items():
        value = ws[cell].value
        if value is None or not str(value).strip():
            updates[field] = default
            continue
        v = str(value).strip().lower()
        if v not in VALID_CHART_TYPES:
            raise SettingsFormError(
                f"'{label}' must be either 'line' or 'bar' (found {value!r})."
            )
        updates[field] = v

    return updates


def _normalize_hex(value: str, label: str) -> str:
    value = value.strip()
    hex_part = value.lstrip("#").upper()
    if len(hex_part) != 6 or any(c not in "0123456789ABCDEF" for c in hex_part):
        raise SettingsFormError(
            f"'{label}' isn't a valid hex color ({value!r}). Use a format like #14495A."
        )
    return "#" + hex_part
