"""
Writes an organization's branding (logo, header text, colors) into a real
copy of the dashboard workbook, using openpyxl.

IMPORTANT -- this module assumes the workbook follows the standard template
layout (same structure every organization's dashboard uses; only the numbers
differ). That layout was confirmed directly against a real dashboard
workbook built from this template on 2026-07-06:

  Sheet "Dashboard":
    A1:K1  merged, solid fill  -> banner primary color + org name text
    A2:K2  merged, solid fill  -> banner secondary color + header subtitle text
    A3:M3  merged, solid fill  -> accent color strip (no text)
    A41    footer text (org name + fiscal year -- left as-is if not provided)
    2 embedded images (header logo + footer logo), both replaced together
    charts, in order:
      0. LineChart "Income"       - 4 series (fiscal years), colors by series index
      1. LineChart "Expense"      - 4 series (fiscal years), colors by series index
      2. BarChart  "Data"          - series 0 = Actual, series 1 = Budgeted

  Sheets "Summary" and "Raw" are never touched -- they hold the org's actual
  financial data/formulas, which is explicitly out of scope for this writer.

If a future template revision changes this layout, update the constants and
assumptions below rather than guessing -- re-inspect the real file first
(see goal prompt, section 6, rule 4: render-and-look, don't assume).

NOTE on display_order (2026-07-07): confirmed by direct save/reload testing
that openpyxl re-sorts a chart's series physically by their `order` value
when the file is saved, and Excel draws series in that same physical
sequence. That means a series' legend position and its draw/z-order are NOT
independently controllable through openpyxl -- setting one moves the other.
So `display_order` intentionally controls both together: it's a single
permutation of the 4 line-chart series, applied by physically reordering
`chart.series` and reassigning idx/order 0-3 to match. Do not attempt to
decouple legend order from z-order without re-testing this assumption first.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import PatternFill

from lib.org_profile import OrgProfile

DASHBOARD_SHEET = "Dashboard"
UNTOUCHED_SHEETS = ("Summary", "Raw")

BANNER_PRIMARY_RANGE = "A1:K1"
BANNER_SECONDARY_RANGE = "A2:K2"
ACCENT_RANGE = "A3:M3"
ORG_NAME_CELL = "A1"
HEADER_SUBTITLE_CELL = "A2"
FOOTER_CELL = "A41"


class TemplateMismatchError(RuntimeError):
    """Raised when the workbook doesn't look like the expected template."""


def _hex(color: str) -> str:
    """openpyxl wants RRGGBB (no '#'); PatternFill additionally wants an AA prefix."""
    return color.lstrip("#").upper()


def _set_fill(ws, cell_range: str, hex_color: str) -> None:
    fill = PatternFill(start_color="FF" + _hex(hex_color), end_color="FF" + _hex(hex_color),
                        fill_type="solid")
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill


def _validate_template(ws) -> None:
    chart_types = [type(c).__name__ for c in ws._charts]
    expected = ["LineChart", "LineChart", "BarChart"]
    if chart_types != expected:
        raise TemplateMismatchError(
            f"Expected charts {expected} on '{DASHBOARD_SHEET}', found {chart_types}. "
            "This workbook doesn't match the standard template -- stopping rather "
            "than guessing which chart is which."
        )
    if len(ws._images) < 1:
        raise TemplateMismatchError(
            f"Expected at least one embedded logo image on '{DASHBOARD_SHEET}', found none."
        )


def _series_props(color: str, is_line: bool) -> GraphicalProperties:
    """Graphical properties for a series, colored correctly for its chart type:
    a line series colors its stroke; a bar series fills its body."""
    if is_line:
        return GraphicalProperties(ln=LineProperties(solidFill=_hex(color)))
    return GraphicalProperties(solidFill=_hex(color))


def _convert_chart_type(chart, target: str):
    """Return `chart` unchanged if it's already the target type ("line"/"bar"),
    otherwise a new chart of the target type carrying over the same series,
    title, anchor, and value-axis minimum.

    Series objects are shared by both LineChart and BarChart in openpyxl, so
    they move across intact (keeping their data references); only the container
    chart type changes. Verify the result in real Excel before trusting it.
    """
    want_line = target == "line"
    is_line = isinstance(chart, LineChart)
    if want_line == is_line:
        return chart

    new = LineChart() if want_line else BarChart()
    if not want_line:
        new.type = "col"          # vertical bars
        new.grouping = "clustered"
    for series in list(chart.series):
        new.series.append(series)
    new.title = chart.title
    new.anchor = chart.anchor
    try:
        new.y_axis.scaling.min = chart.y_axis.scaling.min
    except Exception:
        pass
    return new


def apply_profile_to_workbook(template_xlsx_path: str | Path, profile: OrgProfile,
                               output_xlsx_path: str | Path,
                               logo_path: str | Path | None = None) -> Path:
    """
    Open template_xlsx_path (an existing dashboard workbook), apply the given
    org profile's branding, and save the result to output_xlsx_path.

    template_xlsx_path is opened read-only in memory and never modified --
    output_xlsx_path is always a new/separate file.
    """
    template_xlsx_path = Path(template_xlsx_path)
    output_xlsx_path = Path(output_xlsx_path)

    wb = openpyxl.load_workbook(template_xlsx_path)
    if DASHBOARD_SHEET not in wb.sheetnames:
        raise TemplateMismatchError(f"No '{DASHBOARD_SHEET}' sheet found in {template_xlsx_path.name}")
    ws = wb[DASHBOARD_SHEET]
    _validate_template(ws)

    # 1. Header text
    ws[ORG_NAME_CELL] = profile.org_name
    ws[HEADER_SUBTITLE_CELL] = profile.header_subtitle
    ws[FOOTER_CELL] = f"{profile.org_name} FY{ws[FOOTER_CELL].value[-2:]}" if ws[FOOTER_CELL].value else profile.org_name

    # 2. Banner colors
    _set_fill(ws, BANNER_PRIMARY_RANGE, profile.banner_primary)
    _set_fill(ws, BANNER_SECONDARY_RANGE, profile.banner_secondary)
    _set_fill(ws, ACCENT_RANGE, profile.accent)

    # 3. Logo (replace both anchored copies, keep their existing position/size)
    if logo_path is not None:
        new_images = []
        for old_img in ws._images:
            new_img = XLImage(str(logo_path))
            new_img.anchor = old_img.anchor
            new_images.append(new_img)
        ws._images = new_images

    # 4. Chart type (line/bar), colors, display order, axis minimums.
    #
    # The input always matches the validated template order
    # [Income LineChart, Expense LineChart, Data BarChart], so we identify the
    # three charts by position: 0=Income, 1=Expense, 2=Data. Each may be
    # converted to the type requested in the profile before colors are applied.
    #
    # NOTE (2026-07-06): chart-type conversion (line<->bar) is applied here but
    # can only be fully verified in real Microsoft Excel -- this project's
    # LibreOffice-based rendering doesn't exercise every Excel chart nuance.
    # Treat a converted workbook as "needs a human look in Excel", the same
    # caveat that applies to display_order (see the goal prompt).
    line_colors = profile.line_colors
    display_order = profile.display_order or list(range(4))
    desired_types = {
        0: profile.income_chart_type,
        1: profile.expense_chart_type,
        2: profile.data_chart_type,
    }

    new_charts = []
    for idx, chart in enumerate(list(ws._charts)):
        target = desired_types.get(idx, "line" if idx < 2 else "bar")
        chart = _convert_chart_type(chart, target)
        final_is_line = isinstance(chart, LineChart)

        if idx in (0, 1):  # Income / Expense: 4 fiscal-year series
            # Reorder series per display_order, then reassign idx/order 0..N-1
            # to match the new physical sequence (see module docstring note).
            original = list(chart.series)
            reordered = [original[i] for i in display_order if i < len(original)]
            for slot, series in enumerate(reordered):
                series.idx = slot
                series.order = slot
                color = line_colors[slot] if slot < len(line_colors) else "888888"
                series.graphicalProperties = _series_props(color, final_is_line)
            chart.series = reordered
            chart.y_axis.scaling.min = profile.line_axis_min
        else:  # Data: series[0] = Actual, series[1] = Budgeted
            colors = [profile.bar_actual_color, profile.bar_budget_color]
            for si in range(min(2, len(chart.series))):
                chart.series[si].graphicalProperties = _series_props(colors[si], final_is_line)
            chart.y_axis.scaling.min = profile.bar_axis_min

        new_charts.append(chart)
    ws._charts = new_charts

    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx_path)
    return output_xlsx_path
